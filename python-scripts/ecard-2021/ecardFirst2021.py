#!/usr/bin/python  


import sys
import subprocess
import importlib.util


packages = ['openpyxl', 'pandas', 'xlsx2csv']
for package_name in packages:
    spec = importlib.util.find_spec(package_name)
    if spec is None:
        print("Installing python packages. Please wait.")
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', package_name]) 

 
import openpyxl
import pandas as pd
import os
import shutil


pathCards = "1st-grading-cards"
tempPath = "temp"
infosExcel = "ecard-infos.xlsx"
cardTemplate = "ecard-template.xlsx"
infosSheets = ["Infos-Card-Male", "Infos-Card-Female", "1st-Summary-Male", "1st-Summary-Female"]
sheetsCard = ["infosMale", "infosFemale", "firstGradesMale", "firstGradesFemale"]
firstGradesDFIndex = [2, 3]
dfList = []


def readSheets(infosExcel, tempPath, infosSheets, dfList, sheetsCard):
    print("Loading", infosExcel)
    os.makedirs(tempPath, exist_ok=True)
    for i in range(0, len(infosSheets)):
        csvName = infosSheets[i] + ".csv"
        xlName = infosSheets[i] + ".xlsx"
        sheetNum = str(i + 2)
        subprocess.run(['xlsx2csv', infosExcel, csvName, '-s', sheetNum], shell=False, capture_output=True)
        dfList.insert(i, "df" + str(i))
        dfList[i] = pd.read_csv(csvName)
        dfList[i] = dfList[i].dropna(axis=0, how='any', thresh=5, subset=None, inplace=False)
        writer = pd.ExcelWriter(os.path.join(tempPath, xlName))
        dfList[i].to_excel(writer, infosSheets[i])
        writer.save()
        wb = openpyxl.load_workbook(os.path.join(tempPath, xlName))
        sheetsCard[i] = wb[infosSheets[i]]
        os.remove(csvName)
    shutil.rmtree(tempPath)

    return sheetsCard
    return dfList



def createCard():
    os.makedirs(pathCards, exist_ok=True)
    for i in range(0, 2):
        for k in range(0, len(dfList[i])):
            studentName = dfList[i].iloc[k,1]
            studentCode = dfList[i].iloc[k,0]
            fileName = studentCode + "-" + studentName + ".xlsx"
            print("Creating file:", fileName)
            template = openpyxl.load_workbook(cardTemplate) 
            infoSheet = template["Infos"]
            gradeSheet = template["Grades"]
            studentRow = k + 2
            selectedInfoRow = copyRange(1, studentRow, 11, studentRow, sheetsCard[i]) 
            pasteInfoRow = pasteRange(1, 2, 11, 2, infoSheet, selectedInfoRow)
            selectedFirstGradeRow = copyRange(1, studentRow, 21, studentRow, sheetsCard[firstGradesDFIndex[i]]) 
            pasteFirstGradeRow = pasteRange(1, 2, 21, 2, gradeSheet, selectedFirstGradeRow)
            template.save(os.path.join(pathCards,fileName))
    

def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected
         
 
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1


if __name__ == '__main__':
    readSheets(infosExcel, tempPath, infosSheets, dfList, sheetsCard)
    createCard()
    print("Done!")